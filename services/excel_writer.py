import openpyxl
from openpyxl.cell.cell import MergedCell

def preparar_planilha(caminho_entrada, qtd_geradoras, qtd_beneficiarias):
    """
    Carrega o modelo e cria as abas necessárias para Geradoras e Beneficiárias,
    duplicando os modelos existentes no arquivo Excel.
    """
    wb = openpyxl.load_workbook(caminho_entrada)
    
    # Preparar Geradoras
    if "UC GERADORA" in wb.sheetnames:
        ws_modelo_ger = wb["UC GERADORA"]
        for i in range(qtd_geradoras):
            nome_aba = f"UC GERADORA {i+1}" if i > 0 else "UC GERADORA"
            if i == 0:
                ws_modelo_ger.title = nome_aba
            else:
                nova = wb.copy_worksheet(ws_modelo_ger)
                nova.title = nome_aba

    # Preparar Beneficiárias
    nome_modelo_benef = next((s for s in wb.sheetnames if "UC BENEF" in s.upper()), None)
    if nome_modelo_benef and qtd_beneficiarias > 0:
        ws_modelo_ben = wb[nome_modelo_benef]
        for i in range(qtd_beneficiarias):
            nome_aba = f"UC BENEF. {i+1}"
            if i == 0:
                ws_modelo_ben.title = nome_aba
            else:
                nova = wb.copy_worksheet(ws_modelo_ben)
                nova.title = nome_aba

    return wb

def safe_write(ws, col, row, value):
    """Escreve em células, tratando corretamente células mescladas."""
    coord = f"{col}{row}"
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                ws[rng.start_cell.coordinate].value = value
                return
    else:
        cell.value = value

def salvar_dados_multiplos(wb, dados_estruturados, grupo):
    """
    Mapeia e escreve os dados nas abas corretas. 
    Para Grupo A: Preenche colunas B, C, D (Consumo) e L, M, N (Demanda).
    """
    
    # Formato para bater com a planilha de dimensionamento (image_6498ac.png)
    mapa_meses_a = {
        "JAN": "jan/25", "FEV": "fev/25", "MAR": "mar/25", "ABR": "abr/25",
        "MAI": "mai/25", "JUN": "jun/25", "JUL": "jul/25", "AGO": "ago/25",
        "SET": "set/25", "OUT": "out/25", "NOV": "nov/25", "DEZ": "dez/25"
    }

    # Mapa para o Grupo B
    mapa_meses_b = {k: k.capitalize() for k in mapa_meses_a.keys()}

    for item in dados_estruturados:
        tipo = item['tipo']
        indice = item['indice']
        faturas = item['dados']

        # Identificação da Aba
        if grupo == "A":
            # Tenta a aba específica de Dimensionamento do Grupo A ou a padrão da UC
            nome_aba = "DIMENSIONAMENTO - GRUPO A" if "DIMENSIONAMENTO - GRUPO A" in wb.sheetnames else f"UC BENEF. {indice}"
            if tipo == 'geradora':
                nome_aba = "UC GERADORA" if indice == 1 else f"UC GERADORA {indice}"
        else:
            nome_aba = "UC GERADORA" if tipo == 'geradora' and indice == 1 else f"UC BENEF. {indice}"

        if nome_aba not in wb.sheetnames:
            continue
        
        ws = wb[nome_aba]

        if grupo == "B":
            # --- LÓGICA GRUPO B ---
            cols_b = {
                'consumo': 'F' if tipo == 'beneficiaria' else 'K',
                'saldo': 'Q' if tipo == 'beneficiaria' else 'P',
                'valor': 'J' if tipo == 'beneficiaria' else 'N'
            }
            for dados in faturas:
                mes_ref = mapa_meses_b.get(dados.get("mes"))
                for row in range(5, 40):
                    if str(ws[f"A{row}"].value).strip() == mes_ref:
                        ws[f"{cols_b['consumo']}{row}"] = dados.get("energia_ativa")
                        ws[f"{cols_b['saldo']}{row}"] = dados.get("saldo")
                        ws[f"{cols_b['valor']}{row}"] = dados.get("valor_fatura")
                        break
        
        else:
            # --- LÓGICA GRUPO A (DETALHADO) ---
            for dados in faturas:
                # 1. Preenchimento do Mês da Fatura Atual
                mes_ref = mapa_meses_a.get(dados.get("mes"))
                if mes_ref:
                    for row in range(5, 20):
                        celula_a = str(ws[f"A{row}"].value).strip().lower()
                        if celula_a == mes_ref.lower():
                            # Consumo: B (Ponta), C (Fora Ponta), D (Reservado)
                            ws[f"B{row}"] = dados.get("c_p")   
                            ws[f"C{row}"] = dados.get("c_fp")  
                            ws[f"D{row}"] = dados.get("c_hr")  
                            # Demanda: L (Ponta), M (Fora Ponta), N (Reservado)
                            ws[f"L{row}"] = dados.get("d_p")   
                            ws[f"M{row}"] = dados.get("d_fp")  
                            ws[f"N{row}"] = dados.get("d_hr")
                            break

                # 2. Preenchimento Automático via Histórico (Retroativo)
                if "historico" in dados:
                    for h in dados["historico"]:
                        mes_h = mapa_meses_a.get(h['mes'])
                        # Não sobrescreve o mês atual da fatura
                        if not mes_h or mes_h.lower() == str(mes_ref).lower():
                            continue
                        
                        for row in range(5, 20):
                            celula_a = str(ws[f"A{row}"].value).strip().lower()
                            if celula_a == mes_h.lower():
                                # Consumo
                                ws[f"B{row}"] = h.get('consumo_p')
                                ws[f"C{row}"] = h.get('consumo_fp')
                                ws[f"D{row}"] = h.get('consumo_hr')
                                # Demanda
                                ws[f"L{row}"] = h.get('demanda_p')
                                ws[f"M{row}"] = h.get('demanda_fp')
                                ws[f"N{row}"] = h.get('demanda_hr')
                                break
                                
    return wb